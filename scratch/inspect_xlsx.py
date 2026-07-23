import openpyxl
import os

files = [
    "/Users/bryanpaul/Dropbox/PaulDropbox/E3/docs/Copy of 10 Year Project History - redacted.xlsx",
    "/Users/bryanpaul/Dropbox/PaulDropbox/E3/docs/Reference Sheets/E3_General/Jason Flowers - E3 Texas School Districts Completed Projects 8.1.17.xlsx"
]

for filepath in files:
    if os.path.exists(filepath):
        print(f"\n--- File: {filepath} ---")
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            print("Sheets:", wb.sheetnames)
            for sheetname in wb.sheetnames[:3]:
                sheet = wb[sheetname]
                print(f"Sheet: {sheetname}")
                rows = list(sheet.iter_rows(max_row=3, values_only=True))
                for i, r in enumerate(rows):
                    print(f"  Row {i+1}: {r[:10]}")
        except Exception as e:
            print("Error loading workbook:", e)
    else:
        print(f"File not found: {filepath}")
